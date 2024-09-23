import re
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import load_workbook

from src.backend.objects.company import Company
from src.backend.objects.country import Country
from src.backend.objects.license_type import LicenseType
from src.backend.objects.order import Order
from src.backend.objects.order_template import OrderTemplate
from src.backend.objects.quarter import Quarter
from src.backend.objects.software import Software
from src.backend.objects.software_class import SoftwareClass
from src.backend.objects.software_class_section import SoftwareClassSection
from src.backend.objects.supervisor import Supervisor
from src.utils.config import AppConfig


def get_order_template(order_template_file_path: str) -> OrderTemplate:
    order_template_workbook = load_workbook(filename=order_template_file_path, data_only=True)
    sheets_data = {}
    sheets_titles = []
    for sheet in order_template_workbook.worksheets:
        sheet_data = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheets_titles.append(sheet.title)
        sheets_data[sheet.title] = sheet_data
    check_mandatory_sheets(order_template_file_path, sheets_titles)
    countries = []
    countries_map = {}
    for row in sheets_data["Страны"][1:]:
        if row[0] is not None:
            country_name = str(row[0])
            country = Country(name=country_name)
            countries.append(country)
            countries_map[country.name] = country
    if len(countries) == 0:
        error_message = "Нет валидных данных в справочнике 'Страны'."
        raise ValueError(error_message)
    companies = []
    companies_map = {}
    for row in sheets_data["Наименование проекта"][1:]:
        if row[0] is not None:
            company_name = str(row[0])
            company = Company(name=company_name)
            companies.append(company)
            companies_map[company_name] = company
    if len(companies) == 0:
        error_message = "Нет валидных данных в справочнике 'Наименование проекта'."
        raise ValueError(error_message)
    supervisors = []
    supervisors_map = {}
    for row in sheets_data["ФИО Руководителя"][1:]:
        if row[0] is not None:
            supervisor_name = str(row[0])
            supervisor_email = str(row[1]) if row[1] is not None else None
            supervisor = Supervisor(name=supervisor_name, email=supervisor_email)
            supervisors.append(supervisor)
            supervisors_map[supervisor_name] = supervisor
    licenses_types = []
    licenses_types_map = {}
    for row in sheets_data["тип лицензии"][1:]:
        if row[0] is not None:
            license_type_name = str(row[0])
            license_type = LicenseType(name=license_type_name)
            licenses_types.append(license_type)
            licenses_types_map[license_type_name] = license_type
    software_classes = []
    software_classes_map = {}
    software_class_section = SoftwareClassSection(name="Дополнительные")
    for row in sheets_data["Дополнительный классификатор"][1:]:
        if row[0] is not None:
            software_class_point = None
            software_class_name = row[0]
            target_indicator_name = None
            target_indicator_current_year_first_half = None
            target_indicator_current_year_second_half = None
            target_indicator_next_year_first_half = None
            target_indicator_next_year_second_half = None

            software_class = SoftwareClass(
                class_section=software_class_section,
                class_point=software_class_point,
                class_name=software_class_name,
                target_indicator_name=target_indicator_name,
                target_indicator_current_year_first_half=target_indicator_current_year_first_half,
                target_indicator_current_year_second_half=target_indicator_current_year_second_half,
                target_indicator_next_year_first_half=target_indicator_next_year_first_half,
                target_indicator_next_year_second_half=target_indicator_next_year_second_half,
            )
            software_classes.append(software_class)
            software_classes_map[software_class.class_name] = software_class

    for row in sheets_data["Классификатор Минкомсвязи"][3:]:
        if row[1] is None and row[2] is None and row[3] is None and row[4] is None and row[5] is None and row[6] is None:
            software_class_section = SoftwareClassSection(name=str(row[0]))
        else:
            software_class_point = str(row[0])
            software_class_name = str(row[1])
            if row[2] != "-":
                target_indicator_name = str(row[2])

                # Is None checking
                target_indicator_current_year_first_half = row[3]
                target_indicator_current_year_second_half = row[4] or target_indicator_current_year_first_half
                target_indicator_next_year_first_half = row[5] or target_indicator_current_year_second_half
                target_indicator_next_year_second_half = row[6] or target_indicator_next_year_first_half

            else:
                target_indicator_name = None

                target_indicator_current_year_first_half = None
                target_indicator_current_year_second_half = None
                target_indicator_next_year_first_half = None
                target_indicator_next_year_second_half = None

            software_classes.append(
                SoftwareClass(
                    class_section=software_class_section,
                    class_point=software_class_point,
                    class_name=software_class_name,
                    target_indicator_name=target_indicator_name,
                    target_indicator_current_year_first_half=target_indicator_current_year_first_half,
                    target_indicator_current_year_second_half=target_indicator_current_year_second_half,
                    target_indicator_next_year_first_half=target_indicator_next_year_first_half,
                    target_indicator_next_year_second_half=target_indicator_next_year_second_half,
                )
            )
    software_list = []
    software_list_map = {}
    for row in sheets_data["Наименование ПО"][1:]:
        if row[0] is not None:
            software_name = str(row[0])
            software_maker_name = str(row[1]) if row[1] is not None else None
            if row[2] in countries_map:
                software_country = countries_map[row[2]]
            else:
                raise ValueError(f"Неизвестная страна: {row[2]}")
            if row[5] in software_classes_map:
                software_software_class = software_classes_map[row[5]]
            else:
                for software_class in software_classes:
                    if re.match(r"\d{1,2}\.\d{1,2}", row[5][0:4]) and row[5][0:4] == software_class.class_name[0:4]:
                        software_software_class = software_class
                        break

            software_website = str(row[3]) if row[3] is not None else None
            software_purpose = str(row[4]) if row[4] is not None else None
            software_software_analogs = str(row[6]) if row[6] is not None else None
            software_company = str(row[7]) if row[7] is not None else None
            software_registry_link = str(row[9]) if row[9] is not None else None
            software_is_in_registry = True if software_registry_link else (row[8].lower() == "да" if row[8] is not None else False)
            software = Software(
                country=software_country,
                software_class=software_software_class,
                name=software_name,
                maker_name=software_maker_name,
                website=software_website,
                purpose=software_purpose,
                software_analogs=software_software_analogs,
                company=software_company,
                registry_link=software_registry_link,
                is_in_registry=software_is_in_registry,
            )
            software_list.append(software)
            software_list_map[software_name] = software
    order_list = []
    for row in sheets_data["Заявки"][3:]:
        if row[3] is not None:
            order_year = int(row[1])
            order_quarter = Quarter(row[2].lower())
            if order_quarter is None:
                raise ValueError(f"Неизвестный квартал: {row[2]}")
            if str(row[3]) in supervisors_map:
                order_supervisor = supervisors_map[str(row[3])]
            else:
                raise ValueError(f"Неизвестный руководитель: {row[3]}")
            order_employee_name = str(row[4])
            if str(row[5]) in software_list_map:
                order_software = software_list_map[str(row[5])]
            else:
                raise ValueError(f"Неизвестное ПО: {row[5]}")
            order_tariff_plan = str(row[7]) if row[7] is not None else None
            order_login_and_password = str(row[8]) if row[8] is not None else None
            order_number_license = int(row[10])
            order_is_new_license = row[11].lower() == "новая"
            order_price_for_one = float(row[12])
            order_licenses_period = row[14].date() if row[14] is not None else None
            if str(row[15]) in licenses_types_map:
                order_license_type = licenses_types_map[str(row[15])]
            else:
                raise ValueError(f"Неизвестный тип лицензии: {row[15]}")
            order_useful_life = str(row[16]).lower()
            if str(row[18]) in companies_map:
                order_company_which_will_use = companies_map[str(row[18])]
            else:
                raise ValueError(f"Неизвестный проект: {row[18]}")
            order_list.append(
                Order(
                    year=order_year,
                    quarter=order_quarter,
                    supervisor=order_supervisor,
                    software=order_software,
                    employee_name=order_employee_name,
                    tariff_plan=order_tariff_plan,
                    login_and_password=order_login_and_password,
                    number_license=order_number_license,
                    is_new_license=order_is_new_license,
                    price_for_one=order_price_for_one,
                    licenses_period=order_licenses_period,
                    license_type=order_license_type,
                    useful_life=order_useful_life,
                    company_which_will_use=order_company_which_will_use,
                )
            )

    return OrderTemplate(
        file_path=order_template_file_path,
        companies=companies,
        countries=countries,
        supervisors=supervisors,
        license_types=licenses_types,
        software_classes=software_classes,
        software_list=software_list,
        order_list=order_list,
    )


def check_mandatory_sheets(order_template_file_path: str, sheets_titles: list[str]) -> None:
    mandatory_sheet_titles = [
        "Заявки",
        "Классификатор Минкомсвязи",
        "Дополнительный классификатор",
        "Наименование ПО",
        "Страны",
        "ФИО Руководителя",
        "Наименование проекта",
        "тип лицензии",
        "квартал",
    ]
    missing_sheets = [title for title in mandatory_sheet_titles if title not in sheets_titles]
    if missing_sheets:
        error_message = f"Файл {order_template_file_path} не содержит следующие необходимые листы: {', '.join(missing_sheets)}"
        raise ValueError(error_message)


def get_main_order_template() -> OrderTemplate:  # TODO: implement compare and edit events by some changes from xlsx file (not relevant)
    main_order_template_path = Path(AppConfig.get_order_path("orders/main_template"))

    order_templates_paths = [str(file) for file in main_order_template_path.glob("*.xlsx")]

    if not order_templates_paths:
        error_message = "Не удалось найти файл с основным шаблоном."
        raise FileNotFoundError(error_message)

    if len(order_templates_paths) != 1:
        error_message = f"Слишком много файлов с основным шаблоном. Найдено {len(order_templates_paths)}."
        raise ValueError(error_message)

    order_template_file_path = order_templates_paths[0]

    return get_order_template(order_template_file_path)


def get_order_templates_paths(start_date: date, end_date: date) -> list[str]:
    order_templates_path = Path(AppConfig.get_order_path("orders/order_templates"))
    # take files, that were last modified in this period
    files = []
    for file in order_templates_path.glob("*.xlsx"):
        file_date = datetime.fromtimestamp(file.stat().st_mtime, tz=UTC).date()
        if start_date <= file_date <= end_date:
            files.append(str(file))

    return files


def get_all_order_template_files() -> list[str]:
    order_templates_path = Path(AppConfig.get_order_path("orders/order_templates"))
    return [str(file) for file in order_templates_path.glob("*.xlsx") if file.is_file() and not file.name.startswith("~$")]


def get_all_order_templates() -> list[OrderTemplate]:
    order_templates_paths = get_all_order_template_files()

    return [get_order_template(order_template_path) for order_template_path in order_templates_paths]


def get_order_templates_by_period(start_date: date, end_date: date) -> list[OrderTemplate]:
    order_templates_paths = get_order_templates_paths(start_date, end_date)

    return [get_order_template(order_template_path) for order_template_path in order_templates_paths]
