from collections.abc import Callable
from operator import attrgetter
from typing import TypeVar

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.backend.objects.grouped_order import GroupedOrder
from src.backend.objects.order import Order, has_analogs

T = TypeVar("T")


def group_orders(orders: list[Order]) -> list[GroupedOrder]:
    """Groups orders into groups based on their supervisor, software, tariff plan and is_new_license."""
    group_dict: dict = {}

    for order in orders:
        key = (
            order.supervisor.name,
            order.software.name,
            order.tariff_plan,
            order.is_new_license,
            order.login_and_password,
            order.licenses_period,
            order.license_type.name,
            order.useful_life,
            order.company_which_will_use.name,
            order.price_for_one,
        )

        if key not in group_dict:
            group_dict[key] = []

        group_dict[key].append(order)

    def verify_order_list(getter: Callable[[Order], T], value: T, order_list):
        return all(getter(order) == value for order in order_list)

    # verify
    for key, order_list in group_dict.items():
        first_order = order_list[0]
        if not verify_order_list(lambda order: order.supervisor.name, key[0], order_list):
            raise ValueError(f"Supervisor name mismatch for group {key}")
        if not verify_order_list(lambda order: order.software.name, key[1], order_list):
            raise ValueError(f"Software name mismatch for group {key}")
        if not verify_order_list(lambda order: order.tariff_plan, key[2], order_list):
            raise ValueError(f"Tariff plan mismatch for group {key}")
        if not verify_order_list(lambda order: order.is_new_license, key[3], order_list):
            raise ValueError(f"Is new license mismatch for group {key}")
        if not verify_order_list(lambda order: order.login_and_password, first_order.login_and_password, order_list):
            raise ValueError(f"Login and password are not the same for all orders in the group {key}")
        if not verify_order_list(lambda order: order.licenses_period, first_order.licenses_period, order_list):
            raise ValueError(f"Licenses period are not the same for all orders in the group {key}")
        if not verify_order_list(lambda order: order.license_type.name, first_order.license_type.name, order_list):
            raise ValueError(f"License type is not the same for all orders in the group {key}")
        if not verify_order_list(lambda order: order.useful_life, first_order.useful_life, order_list):
            raise ValueError(f"Useful life is not the same for all orders in the group {key}")
        if not verify_order_list(lambda order: order.company_which_will_use.name, first_order.company_which_will_use.name, order_list):
            raise ValueError(f"Project is not the same for all orders in the group {key}")
        if not verify_order_list(lambda order: order.price_for_one, first_order.price_for_one, order_list):
            raise ValueError(f"Price for one is not the same for all orders in the group {key}")

    grouped_orders = []
    for order_list in group_dict.values():
        supervisor = order_list[0].supervisor
        employee_names = ", ".join([order.employee_name for order in order_list])
        software = order_list[0].software
        tariff_plan = order_list[0].tariff_plan
        login_and_password = order_list[0].login_and_password
        number_orders = len(order_list)
        number_license = sum([order.number_license for order in order_list])
        price_for_one = order_list[0].price_for_one
        licenses_period = order_list[0].licenses_period
        license_type = order_list[0].license_type
        useful_life = order_list[0].useful_life
        company_which_will_use = order_list[0].company_which_will_use
        is_new_license = order_list[0].is_new_license

        grouped_order = GroupedOrder(
            supervisor=supervisor,
            employee_names=employee_names,
            software=software,
            tariff_plan=tariff_plan,
            login_and_password=login_and_password,
            number_orders=number_orders,
            number_license=number_license,
            price_for_one=price_for_one,
            licenses_period=licenses_period,
            license_type=license_type,
            useful_life=useful_life,
            is_new_license=is_new_license,
            company_which_will_use=company_which_will_use,
        )

        grouped_orders.append(grouped_order)

    return grouped_orders


def create_orders_report_with_sort_by_params(order_list: list[Order] | None, sort_params: dict[str, bool]) -> Workbook | None:
    if order_list is None:
        return None

    grouped_orders = group_orders(order_list)

    params = list(sort_params.keys())
    sort_orders = list(sort_params.values())

    for i in range(len(params) - 1, -1, -1):
        param = params[i]
        reverse_order = not sort_orders[i]
        grouped_orders = sorted(grouped_orders, key=attrgetter(param), reverse=reverse_order)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"

    headers = [
        "№",
        "ФИО руководителя согласовавшего закупку",
        "ФИО сотрудников, для которых приобретается ПО",
        "Наименование ПО",
        "Производитель ПО",
        "Тарифный план (если несколько вариантов)",
        """Логин и пароль
(если есть уже аккаунт, если нет, то мы сами создадим)""",
        "Страна производства",
        "Общее количество лицензий",
        "Продление/новая",
        "Цена за единицу, руб, без НДС",
        "Стоимость, руб, без НДС (за все лицензии)",
        """Окончание срока действия текущих лицензий 
(если продление)""",
        """Лицензия/
Подписка/
Техподдержка (ТП)/
Лицензия+ТП/сертификат, домен""",
        """Срок полезного использования
(месяц/
год/
бессрочно)""",
        "Тип ПО по классификатору Минкомсвязи",
        """Наличие ПО в реестре Российского ПО*
(да/нет)""",
        """Наличие аналогов в Реестре Российского ПО 
(да/нет)""",
        "Проект в котором будет использоваться ПО",
        "Ссылка на ПО (лицензии) на сайте производителя",
        "Альтернативы",
    ]

    left_right_bold_border = Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="thin"), bottom=Side(style="thin"))

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    default_font = Font(name="Calibri", size=12)
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    header_fill = PatternFill(start_color="70ad47", end_color="70ad47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    sheet.append(headers)
    sheet.row_dimensions[1].height = 120

    for col_num, _ in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = left_right_bold_border

    special_fill = PatternFill(start_color="96BD7C", end_color="96BD7C", fill_type="solid")

    special_columns = [
        "Производитель ПО",
        "Страна производства",
        "Тип ПО по классификатору Минкомсвязи",
        """Наличие ПО в реестре Российского ПО*
(да/нет)""",
        """Наличие аналогов в Реестре Российского ПО 
(да/нет)""",
        "Проект в котором будет использоваться ПО",
        "Ссылка на ПО (лицензии) на сайте производителя",
        "Альтернативы",
    ]

    for header in special_columns:
        col_idx = headers.index(header) + 1
        column_letter = get_column_letter(col_idx)
        for row in sheet[column_letter]:
            row.fill = special_fill
            row.border = thin_border
    column_widths = {
        "A": 5,  # №
        "B": 25,  # ФИО руководителя согласовавшего закупку
        "C": 25,  # ФИО сотрудников, для которых приобретается ПО
        "D": 20,  # Наименование ПО
        "E": 20,  # Производитель ПО
        "F": 20,  # Тарифный план
        "G": 30,  # Логин и пароль
        "H": 15,  # Страна производства
        "I": 10,  # Общее количество лицензий
        "J": 15,  # Продление/новая
        "K": 20,  # Цена за единицу
        "L": 25,  # Стоимость за все лицензии
        "M": 25,  # Окончание срока действия лицензий
        "N": 35,  # Лицензия/Подписка/Техподдержка
        "O": 20,  # Срок полезного использования
        "P": 30,  # Тип ПО по классификатору
        "Q": 15,  # Наличие ПО в реестре
        "R": 20,  # Наличие аналогов в реестре
        "S": 30,  # Проект
        "T": 30,  # Ссылка на ПО
        "U": 30,  # Альтернативы
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    for index, order in enumerate(grouped_orders):
        row = [
            index + 1,
            order.supervisor.name,
            order.employee_names,
            order.software.name,
            order.software.maker_name if order.software.maker_name is not None else "",
            order.tariff_plan if order.tariff_plan is not None else "",
            order.login_and_password if order.login_and_password is not None else "",
            order.software.country.name,
            order.number_license,
            "Новая" if order.is_new_license else "Продление",
            order.price_for_one,
            order.price_for_one * order.number_license,
            order.licenses_period.strftime("%Y.%m.%d") if order.licenses_period is not None else "",
            order.license_type.name,
            order.useful_life,
            order.software.software_class.class_name,
            "Да" if order.software.is_in_registry else "Нет",
            "Да" if has_analogs(order.software.software_analogs) else "Нет",
            order.company_which_will_use.name,
            order.software.website if order.software.website is not None else "",
            order.software.software_analogs if order.software.software_analogs is not None else "",
        ]
        sheet.append(row)

        for col_num, _ in enumerate(row, 1):
            cell = sheet.cell(row=index + 2, column=col_num)
            cell.font = default_font
            cell.border = thin_border

    return workbook
